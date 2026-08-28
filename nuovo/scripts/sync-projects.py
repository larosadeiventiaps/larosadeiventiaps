"""
Sincronizza projects.json dal file Excel "Lista Progetti.xlsx".

Uso:
  python sync-projects.py <percorso-excel>

Se non viene passato un percorso, usa il file predefinito su OneDrive.

La colonna "Immagini" dell'Excel deve contenere il nome del file immagine
(es. "genitori-a-confronto.jpg"). Le immagini vanno messe nella cartella
new-site/images/projects/.
"""

import json
import os
import re
import sys
from datetime import datetime, date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(SITE_DIR, "data", "projects.json")

DEFAULT_EXCEL = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Progetti",
    "Lista Progetti.xlsx",
)

# Mappa colonne Excel -> campi JSON
COL_MAP = {
    0: "startDate",
    1: "endDate",
    2: "title",
    3: "description",
    4: "incontri",
    5: "oreIncontro",
    6: "partecipanti",
    7: "educatori",
    8: "volontari",
    9: "sponsor",
    10: "professionisti",
    11: "collaboratori",
    12: "ore",
    13: "image",
}

MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}



# ---------------------------------------------------------------------------
# I corsi che hanno cambiato nome, e l'Excel non lo sa
# ---------------------------------------------------------------------------
# ⛔ **Un corso rinominato dopo la prima edizione.** Il titolare, il 28/08/2026:
# «il primo progetto si chiama "Musica ed espressione Corporea" con una sola
# edizione; tutti gli altri sono "Musica e Movimento" con tutte le varie
# edizioni». Il foglio Excel storico porta ancora il nome vecchio su tutte le
# righe, e non lo si riscrive: e' l'originale, e serve a ritrovare da dove
# viene un dato il giorno che un accostamento si rivelasse sbagliato — la
# stessa ragione per cui `ProgettoPartner.nomeOriginale` esiste nel gestionale.
#
# ⚠️ **La chiave e' (titolo, data d'inizio), non il solo titolo**: rinominare
# per titolo cambierebbe anche la prima edizione, che il nome vecchio deve
# tenerlo. E' la stessa trappola gia' pagata il 24/08 separando i progetti
# dagli eventi — «Prim'Olio» esiste in tre anni, e per titolo si cancellavano
# anche le edizioni sbagliate.
#
# ⚠️ **Senza questa tabella la rigenerazione non "sbaglia il nome": perde i
# dati.** Il riaggancio di `appuntamenti` e `luogoId` piu' sotto cerca la riga
# vecchia per (titolo, data d'inizio); con il titolo rimesso a quello
# dell'Excel, le righe gia' rinominate in `projects.json` non si trovano piu',
# e quel che gli era attaccato resta indietro. L'avviso lo direbbe, ma dopo.
RINOMINATI = {
    # (titolo nell'Excel, data d'inizio): titolo giusto
    ("Musica ed Espressione corporea", "2025-09-27"): "Musica e Movimento",
    ("Musica ed Espressione corporea", "2026-02-07"): "Musica e Movimento",
    ("Musica ed Espressione corporea", "2026-09-19"): "Musica e Movimento",
}


def rinomina(titolo, data_inizio):
    """Il nome giusto per questa riga, o quello che aveva se non e' cambiato."""
    return RINOMINATI.get((titolo, data_inizio), titolo)


def risolvi_sponsor(grezzo):
    """I nomi degli sponsor, risolti con la stessa tabella di sync-partners.py.

    ⛔ La tabella NON si copia qui: se un giorno qualcuno correggesse un nome
    in un solo file, la pagina Partner e la pagina Progetti chiamerebbero lo
    stesso ente in due modi diversi, e nessuno saprebbe quale dei due è giusto.
    """
    import importlib.util
    percorso = os.path.join(SCRIPT_DIR, "sync-partners.py")
    spec = importlib.util.spec_from_file_location("sync_partners", percorso)
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo.resolve_partner_names(grezzo)


def parse_date(val):
    """Converte un valore cella in una data ISO (YYYY-MM-DD)."""
    if val is None:
        return None

    # Già un oggetto datetime/date
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()

    s = str(val).strip()
    if not s:
        return None

    # Formato "7 luglio 2024" o "21 marzo 2025"
    match = re.match(r"(\d{1,2})\s+(\w+)\s+(\d{4})", s)
    if match:
        day, month_name, year = match.groups()
        month = MESI_IT.get(month_name.lower())
        if month:
            return f"{year}-{month:02d}-{int(day):02d}"

    # Formato "gennaio 2023" (senza giorno)
    match = re.match(r"(\w+)\s+(\d{4})", s)
    if match:
        month_name, year = match.groups()
        month = MESI_IT.get(month_name.lower())
        if month:
            return f"{year}-{month:02d}-01"

    # Formato "26/11/2020"
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if match:
        day, month, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"

    # Anno solo "2023"
    match = re.match(r"^(\d{4})$", s)
    if match:
        return f"{s}-01-01"

    return None


def determine_status(start_str, end_str):
    """Determina lo stato del progetto in base alle date."""
    today = date.today()

    start = None
    end = None
    if start_str:
        try:
            start = date.fromisoformat(start_str)
        except ValueError:
            pass
    if end_str:
        try:
            end = date.fromisoformat(end_str)
        except ValueError:
            pass

    if start and start > today:
        return "futuro"
    if end and end < today:
        return "passato"
    if start and start <= today:
        return "in_corso"

    return "passato"


def to_int(val):
    """Converte in intero se possibile."""
    if val is None:
        return None
    try:
        n = int(float(val))
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def to_float(val):
    """Converte in float se possibile."""
    if val is None:
        return None
    try:
        n = float(val)
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def build_description(project):
    """Genera una descrizione dal titolo e dai collaboratori/sponsor."""
    parts = []
    title = project.get("title", "")

    if project.get("collaboratori"):
        parts.append(f"In collaborazione con {project['collaboratori']}.")
    if project.get("sponsor"):
        parts.append(f"Con il sostegno di {project['sponsor']}.")
    if project.get("incontri") and project.get("ore"):
        parts.append(f"{project['incontri']} incontri per un totale di {project['ore']} ore.")
    elif project.get("incontri"):
        parts.append(f"{project['incontri']} incontri.")
    if project.get("partecipanti"):
        parts.append(f"{project['partecipanti']} partecipanti.")

    return " ".join(parts) if parts else title


# Campi che l'Excel non conosce e che nel frattempo qualcuno può aver
# scritto a mano in projects.json (gli appuntamenti con l'ora, il luogo
# dell'anagrafica, il vecchio luogo in chiaro): questo script li riscrive
# da capo a ogni sincronizzazione, e senza questa protezione una
# rigenerazione li cancellerebbe in silenzio (26/08/2026).
CAMPI_DA_CONSERVARE = ("appuntamenti", "luogoId", "location")


def carica_progetti_esistenti():
    """I progetti già in projects.json, indicizzati per (titolo, data inizio).

    ⛔ Si accoppia per titolo **e** data d'inizio, mai per il solo titolo:
    "Teatro" o "Cinema" tornano ogni anno con lo stesso nome, e agganciare
    gli appuntamenti del laboratorio 2026 a quello del 2025 sarebbe un dato
    sbagliato, non solo un dato perso.
    """
    if not os.path.exists(OUTPUT_PATH):
        return {}
    try:
        with open(OUTPUT_PATH, "r", encoding="utf-8") as f:
            vecchi = json.load(f)
    except (OSError, ValueError):
        return {}
    indice = {}
    for p in vecchi:
        chiave = (p.get("title"), p.get("startDate"))
        if chiave[0] is None or chiave[1] is None:
            continue
        conservati = {k: p[k] for k in CAMPI_DA_CONSERVARE if k in p}
        if conservati:
            indice[chiave] = conservati
    return indice


def sync(excel_path):
    """Legge l'Excel e genera projects.json."""
    try:
        import openpyxl
    except ImportError:
        print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(excel_path):
        print(f"Errore: file non trovato: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    progetti_esistenti = carica_progetti_esistenti()
    chiavi_riagganciate = set()

    projects = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Salta righe vuote o righe di riepilogo (colonna 14/15 con "Totale...")
        if not row or len(row) < 3:
            continue

        title = row[2]
        if not title or str(title).strip() == "":
            continue

        title = str(title).strip()

        # Salta se sembra una riga di riepilogo
        if title.startswith("Totale") or title.startswith("Nr."):
            continue

        start_raw = parse_date(row[0])
        end_raw = parse_date(row[1]) if len(row) > 1 else None

        # Se non c'è data fine, usa data inizio
        if not end_raw and start_raw:
            end_raw = start_raw

        # Qui, e non prima: la chiave di `RINOMINATI` e' (titolo, data
        # d'inizio), quindi il nome giusto si sa solo dopo aver letto la data.
        # Risolvendolo in questo punto vale sia per la riga che si scrive sia
        # per `chiave = (title, start_raw)` del riaggancio piu' sotto - un
        # posto solo, che e' il motivo per cui non si risolve in due punti.
        title = rinomina(title, start_raw)

        status = determine_status(start_raw, end_raw)

        # Colonna 3 = Descrizione Progetto (dall'Excel)
        excel_description = str(row[3]).strip() if len(row) > 3 and row[3] else None

        # Colonna 13 = Immagine
        img_val = row[13] if len(row) > 13 else None
        if img_val and str(img_val).strip() and str(img_val).strip().lower() != "x":
            image = f"images/projects/{str(img_val).strip()}"
        else:
            image = "images/logo.jpg"

        project = {"title": title}

        # Campi numerici (colonne 4-12, la 3 è descrizione)
        incontri = to_int(row[4]) if len(row) > 4 else None
        ore_incontro = to_float(row[5]) if len(row) > 5 else None
        partecipanti = to_int(row[6]) if len(row) > 6 else None
        educatori = to_int(row[7]) if len(row) > 7 else None
        volontari = to_int(row[8]) if len(row) > 8 else None
        sponsor = str(row[9]).strip() if len(row) > 9 and row[9] else None
        professionisti = to_int(row[10]) if len(row) > 10 else None
        collaboratori = str(row[11]).strip() if len(row) > 11 and row[11] else None
        ore_totali = to_float(row[12]) if len(row) > 12 else None

        project["image"] = image
        if start_raw:
            project["startDate"] = start_raw
        if end_raw:
            project["endDate"] = end_raw
        project["status"] = status

        if incontri:
            project["incontri"] = incontri
        if ore_totali and ore_totali > 0:
            project["ore"] = ore_totali
        if partecipanti:
            project["partecipanti"] = partecipanti
        if educatori:
            project["educatori"] = educatori
        if volontari:
            project["volontari"] = volontari
        if professionisti:
            project["professionisti"] = professionisti
        if sponsor:
            # ⭐ `sponsor` è un ELENCO di nomi che combaciano con partners.json,
            #    non la stringa grezza del foglio (24/08/2026). Il foglio scrive
            #    «Comune BaR?», «Scuola Redi - IC Caponnetto», e celle con due
            #    enti dentro: quei nomi finiscono su una pagina pubblica accanto
            #    a «Con il sostegno di». La tabella che li risolve è una sola,
            #    in `sync-partners.py`, e sta lì perché serve a entrambi:
            #    duplicarla vorrebbe dire due elenchi liberi di divergere.
            project["sponsor"] = risolvi_sponsor(sponsor)
        if collaboratori:
            project["collaboratori"] = collaboratori

        # Usa la descrizione dall'Excel se presente, altrimenti genera automatica
        if excel_description:
            project["description"] = excel_description
        else:
            project["description"] = build_description(project)

        # Riaggancia appuntamenti/luogoId/location scritti a mano nella riga
        # già esistente con lo stesso titolo e la stessa data d'inizio: se
        # non lo facciamo qui, la riga sotto li perde senza dirlo a nessuno.
        chiave = (title, start_raw)
        if chiave in progetti_esistenti:
            for campo, valore in progetti_esistenti[chiave].items():
                project[campo] = valore
            chiavi_riagganciate.add(chiave)

        projects.append(project)

    # In corso e futuri: dal più vicino al più lontano (crescente)
    # Passati: dal più recente al più vecchio (decrescente)
    in_corso = sorted([p for p in projects if p["status"] == "in_corso"],
                      key=lambda p: p.get("startDate", "9999"))
    futuri = sorted([p for p in projects if p["status"] == "futuro"],
                    key=lambda p: p.get("startDate", "9999"))
    passati = sorted([p for p in projects if p["status"] == "passato"],
                     key=lambda p: p.get("startDate", "0000"), reverse=True)
    projects = in_corso + futuri + passati

    # Salva JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)

    # Riepilogo
    tot = len(projects)
    in_corso = sum(1 for p in projects if p["status"] == "in_corso")
    futuro = sum(1 for p in projects if p["status"] == "futuro")
    passato = sum(1 for p in projects if p["status"] == "passato")
    tot_ore = sum(p.get("ore", 0) for p in projects)
    tot_part = sum(p.get("partecipanti", 0) for p in projects)
    tot_incontri = sum(p.get("incontri", 0) for p in projects)

    print(f"Sincronizzazione completata!")
    print(f"  {tot} progetti ({in_corso} in corso, {futuro} futuri, {passato} passati)")
    print(f"  {tot_incontri} incontri, {tot_ore} ore, {tot_part} partecipanti")
    print(f"  Salvato in: {OUTPUT_PATH}")

    # Appuntamenti/luogoId/location scritti a mano che c'erano nel vecchio
    # projects.json e che questa sincronizzazione NON è riuscita a
    # riagganciare (titolo o data d'inizio cambiati nell'Excel): sono andati
    # persi con questo salvataggio, e va saputo subito, non scoperto dopo.
    perse = [chiave for chiave in progetti_esistenti if chiave not in chiavi_riagganciate]
    if perse:
        print(f"\n  ATTENZIONE — dati scritti a mano NON riagganciati (persi in questo salvataggio):")
        for titolo, data_inizio in perse:
            campi = ", ".join(progetti_esistenti[(titolo, data_inizio)].keys())
            print(f"    - \"{titolo}\" ({data_inizio}): {campi}")
        print(f"    Titolo o data d'inizio sono cambiati nell'Excel rispetto a {OUTPUT_PATH}:")
        print(f"    controlla a mano prima di ridare questi dati.")
    elif progetti_esistenti:
        print(f"\n  {len(chiavi_riagganciate)} riga/e con appuntamenti/luogoId/location scritti a mano riagganciate.")

    # Controlla immagini mancanti
    missing = []
    for p in projects:
        img_path = p["image"]
        if img_path == "images/logo.jpg":
            continue  # fallback al logo, sempre presente
        full_path = os.path.join(SITE_DIR, img_path)
        if not os.path.exists(full_path):
            missing.append(img_path)
    if missing:
        print(f"\n  ATTENZIONE — Immagini mancanti in images/projects/:")
        for m in missing:
            print(f"    - {m}")


# ---------------------------------------------------------------------------
# ⛔⛔ CHIUSO A CHIAVE — il sito non si rigenera piu' dall'Excel (28/08/2026)
# ---------------------------------------------------------------------------
# Il titolare: «puoi bloccare l'aggiornamento automatico del sito dai file di
# excel. Bisogna fare in modo che il sito si aggiorni dal gestionale».
#
# ⚠️ **Non e' una precauzione teorica: questo script DISTRUGGE lavoro vero.**
# I file in `data/` portano ormai cose che nell'Excel non ci sono e non ci
# possono stare — le date dei singoli incontri, i luoghi con l'indirizzo, i
# loghi dei partner, le descrizioni, i nomi corretti dei progetti rinominati.
# Una rigenerazione le riportava tutte a stringa vuota **senza un errore e
# senza un avviso**: il sito si svuotava, e lo si scopriva guardandolo.
#
# ⛔ **Non si cancella lo script.** L'Excel resta la fonte storica del lavoro
# di sei anni, e il giorno in cui quei dati si travaseranno nel gestionale
# questo codice serve a rileggerlo. Chi lo lancia oggi, pero', quasi sempre non
# sa che cosa sta per perdere.
#
# ⇒ Per usarlo davvero: `--forza-so-cosa-sto-facendo`, e prima si fa una copia
#   di `data/`. Il nome dell'interruttore e' lungo apposta.
def _fermati_se_non_forzato():
    import sys as _sys
    if "--forza-so-cosa-sto-facendo" in _sys.argv:
        _sys.argv.remove("--forza-so-cosa-sto-facendo")
        print("\u26a0\ufe0f  Rigenerazione FORZATA: quel che era scritto a mano e a rischio.")
        return
    print(__doc__ or "")
    print("\u26d4 Questo script e' CHIUSO A CHIAVE dal 28/08/2026.")
    print()
    print("   Il sito non si rigenera piu' dall'Excel: i file in data/ contengono")
    print("   ormai dati che nell'Excel non esistono (appuntamenti, luoghi con")
    print("   indirizzo, loghi, descrizioni, nomi corretti), e rigenerare li")
    print("   cancella in silenzio.")
    print()
    print("   La strada nuova e' il gestionale: vedi docs/ e la memoria di")
    print("   progetto. Se DEVI davvero rigenerare, fai una copia di data/ e")
    print("   rilancia con  --forza-so-cosa-sto-facendo")
    raise SystemExit(2)

if __name__ == "__main__":
    _fermati_se_non_forzato()
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    sync(excel_path)
