"""
Sincronizza partners.json dal file Excel "Partner.xlsx" e/o dalla lista progetti.

Uso:
  python sync-partners.py [--partner-excel <percorso>] [--projects-excel <percorso>]

Senza argomenti usa i file predefiniti su OneDrive.

Il file Partner.xlsx ha le colonne:
  Nome | Tipo | Descrizione | Logo | Sito Web

Il file Lista Progetti.xlsx viene usato per estrarre nomi partner
dai campi sponsor e collaboratori. Eventuali nuovi partner trovati
vengono aggiunti con dati placeholder.
"""

import json
import os
import re
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(SITE_DIR, "data", "partners.json")

ONEDRIVE_BASE = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Progetti",
)

DEFAULT_PARTNER_EXCEL = os.path.join(ONEDRIVE_BASE, "Partner.xlsx")
DEFAULT_PROJECTS_EXCEL = os.path.join(ONEDRIVE_BASE, "Lista Progetti.xlsx")


# Mappa abbreviazioni/varianti usate nell'Excel progetti → nome completo
# nel Partner.xlsx. I nomi qui sotto vengono ignorati quando appaiono
# nell'Excel progetti perché sono già coperti dal partner corretto.
ALIASES = {
    "comune bar": "Comune di Bagno a Ripoli",
    "bar": "Comune di Bagno a Ripoli",
    "fondazione cr": "Fondazione CR Firenze",
    "cri": "Croce Rossa Italiana",
    "biblioteca": "Biblioteca di Bagno a Ripoli",
    "scuola redi": "Scuola Francesco Redi",
    "elsa morante": "Istituto Elsa Morante",
    "cdp grassina": "Casa del Popolo Grassina",
    "cdp balatro": "Casa del Popolo Balatro",
    "contrada alfiere": "Contrada dell'Alfiere",
    "becare": "BE Care S.r.l.",
    "mise antella": "Misericordia di Antella",
    # ⛔ **Non e' la Misericordia: a Ponte a Ema la pubblica assistenza e' la
    # CROCE D'ORO** (titolare, 28/08/2026). Nel foglio storico la sigla e'
    # scritta «mise ponte a ema» per simmetria con «mise antella» e «mise
    # badia», che invece sono Misericordie vere. ⚠️ La sigla nell'Excel NON si
    # corregge: e' l'originale, e questa tabella esiste proprio per tradurre
    # le grafie storiche senza riscrivere la fonte.
    "mise ponte a ema": "Croce d'Oro Ponte a Ema",
    "mise badia": "Misericordia di Badia a Ripoli",
    "fpgrassina": "Fratellanza Popolare Grassina",
    "croce d'oro ponte a ema": "Croce d'Oro Ponte a Ema",
    "l'apiario": "Apicoltura San Martino",
    "fontenuova": "Cooperativa Fontenuova",
    # Aggiunti il 24/08/2026: sono le grafie che compaiono nella colonna
    # «sponsor» del foglio progetti, diverse dalle sigle brevi qui sopra.
    # Servono perché quei nomi finiscono su una pagina pubblica accanto a
    # «Con il sostegno di»: sono enti che ci hanno dato dei soldi, e
    # «Comune BaR?» col punto interrogativo è la cosa peggiore che possa
    # capitare a quella riga.
    "comune bar?": "Comune di Bagno a Ripoli",
    "scuola redi - ic caponnetto": "Scuola Francesco Redi",
    "contrada alfiere - bagno a ripoli": "Contrada dell'Alfiere",
    "be care srl": "BE Care Srl",
    "cri bagno a ripoli": "Croce Rossa Italiana",
}

# Stringhe composte nell'Excel che vanno espanse in più partner singoli
COMPOSITE_EXPANSIONS = {
    "cri mise antella fpgrassina mise ponte a ema mise badia": [
        "Croce Rossa Italiana",
        "Misericordia di Antella",
        "Fratellanza Popolare Grassina",
        "Croce d'Oro Ponte a Ema",
        "Misericordia di Badia a Ripoli",
    ],
    "sds firenze sud-est e comune bar": [
        "SdS Firenze Sud-Est",
        "Comune di Bagno a Ripoli",
    ],
    "comune bar il teatro dell'inutile e l'apiario": [
        "Comune di Bagno a Ripoli",
        "Il Teatro dell'Inutile",
        "Apicoltura San Martino",
    ],
    "la lanterna scuola redi": [
        "La Lanterna",
        "Scuola Francesco Redi",
    ],
    "scuola redi fondazione claudio ciai": [
        "Scuola Francesco Redi",
        "Fondazione Claudio Ciai",
    ],
    "becare e cri": [
        "BE Care S.r.l.",
        "Croce Rossa Italiana",
    ],
    "auser legambiente": [
        "Auser",
        "Legambiente",
    ],
    "consorzio blu ancora": [
        "Consorzio Blu",
    ],
    "elsa morante e croce d'oro ponte a ema": [
        "Istituto Elsa Morante",
        "Croce d'Oro Ponte a Ema",
    ],
}


def normalize_name(name):
    """Normalizza un nome partner per confronto."""
    return re.sub(r"\s+", " ", name.strip().lower())


def resolve_partner_names(raw):
    """Risolve una stringa dall'Excel progetti in nomi partner reali."""
    if not raw or not str(raw).strip():
        return []
    s = str(raw).strip()
    key = normalize_name(s)

    # Controlla se è una stringa composita nota
    if key in COMPOSITE_EXPANSIONS:
        return COMPOSITE_EXPANSIONS[key]

    # Controlla se è un alias noto
    if key in ALIASES:
        return [ALIASES[key]]

    # Prova a splittare su separatori semplici (" e ", ", ")
    parts = re.split(r"\s+e\s+|,\s*|/\s*", s)
    resolved = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        pk = normalize_name(p)
        if pk in ALIASES:
            resolved.append(ALIASES[pk])
        else:
            resolved.append(p)
    return resolved


def read_partner_excel(path):
    """Legge Partner.xlsx e restituisce una lista di dict."""
    try:
        import openpyxl
    except ImportError:
        print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(path):
        print(f"Partner.xlsx non trovato: {path}")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    partners = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue

        name = str(row[0]).strip()
        partner = {
            "name": name,
            "type": str(row[1]).strip() if len(row) > 1 and row[1] else "associazione",
            "description": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "logo": "",
            "url": "",
        }

        if len(row) > 3 and row[3] and str(row[3]).strip():
            logo_file = str(row[3]).strip()
            partner["logo"] = f"images/partners/{logo_file}"

        if len(row) > 4 and row[4] and str(row[4]).strip():
            partner["url"] = str(row[4]).strip()

        partners.append(partner)

    return partners


def extract_from_projects(path):
    """Estrae nomi partner unici dall'Excel dei progetti."""
    try:
        import openpyxl
    except ImportError:
        return set()

    if not os.path.exists(path):
        print(f"Lista Progetti.xlsx non trovato: {path}")
        return set()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        title = row[2]
        if not title or str(title).strip() in ("", "Totale", "Nr."):
            continue

        # Colonna 8 = sponsor, colonna 10 = collaboratori
        sponsor = row[8] if len(row) > 8 else None
        collaboratori = row[10] if len(row) > 10 else None

        for raw in [sponsor, collaboratori]:
            for name in resolve_partner_names(raw):
                names.add(name)

    return names


def conserva_lavoro_a_mano(partners):
    """Rimette descrizione, logo e url gia' scritti a mano in `partners.json`.

    ⛔ **Senza questa funzione la rigenerazione SVUOTA la pagina Partner.**
    `read_partner_excel` legge logo e url dall'Excel (colonne 3 e 4), che per
    quasi tutti sono vuote, e la **descrizione nell'Excel non c'e' affatto**: i
    partner che arrivano dal foglio progetti nascono con tutti e tre i campi a
    stringa vuota. Chi rilanciava lo script si ritrovava 32 schede senza logo e
    senza una riga di testo, e nulla glielo diceva — nessun errore, nessun
    avviso, solo un sito improvvisamente vuoto.

    ⚠️ **Vince sempre il valore gia' presente, non quello dell'Excel.** Il JSON
    e' il posto dove quel lavoro e' stato fatto; l'Excel e' un elenco di nomi. Se
    un giorno si vorra' governare i loghi dal foglio, questa regola va rovesciata
    **di proposito**, non per caso.

    ⚠️ L'accostamento e' per nome normalizzato: un ente rinominato nell'Excel
    perde quel che aveva, e **lo si dice** invece di lasciarlo sparire in
    silenzio. E' la stessa lezione di `sync-projects.py`, dove il riaggancio
    degli appuntamenti si fa per (titolo, data d'inizio).
    """
    try:
        with open(OUTPUT_PATH, encoding="utf-8") as f:
            vecchi = json.load(f)
    except (FileNotFoundError, ValueError):
        return partners, []

    per_nome = {normalize_name(v.get("name", "")): v for v in vecchi}
    ripresi = 0
    for p in partners:
        v = per_nome.pop(normalize_name(p["name"]), None)
        if not v:
            continue
        for campo in ("description", "logo", "url"):
            if not p.get(campo) and v.get(campo):
                p[campo] = v[campo]
                ripresi += 1

    # Chi c'era nel JSON e non torna dall'Excel: non lo si cancella di nascosto.
    persi = sorted(v.get("name", "?") for v in per_nome.values()
                   if v.get("description") or v.get("logo") or v.get("url"))
    print(f"  conservati {ripresi} campi scritti a mano")
    return partners, persi


def sync(partner_excel=None, projects_excel=None):
    """Sincronizza partners.json."""
    partner_path = partner_excel or DEFAULT_PARTNER_EXCEL
    projects_path = projects_excel or DEFAULT_PROJECTS_EXCEL

    # 1. Leggi partner da Excel dedicato
    partners = read_partner_excel(partner_path)
    known_names = {normalize_name(p["name"]) for p in partners}

    # 2. Estrai nomi da Excel progetti
    project_names = extract_from_projects(projects_path)

    # 3. Aggiungi partner nuovi (non già presenti)
    added = 0
    for name in sorted(project_names):
        if normalize_name(name) not in known_names:
            partners.append({
                "name": name,
                "type": "associazione",
                "description": "",
                "logo": "",
                "url": "",
            })
            known_names.add(normalize_name(name))
            added += 1

    # 4. Rimetti il lavoro fatto a mano PRIMA di salvare (vedi la funzione)
    partners, persi = conserva_lavoro_a_mano(partners)
    if persi:
        print("  \u26a0\ufe0f  ATTENZIONE: questi partner avevano descrizione/logo/url nel JSON")
        print("      e non tornano dall'Excel. Rinominati? Tolti? Vanno guardati:")
        for n in persi:
            print(f"        - {n}")

    # 5. Ordina per nome
    partners.sort(key=lambda p: p["name"].lower())

    # 5. Salva JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(partners, f, ensure_ascii=False, indent=2)

    print(f"Sincronizzazione partner completata!")
    print(f"  {len(partners)} partner totali ({added} nuovi da progetti)")
    print(f"  Salvato in: {OUTPUT_PATH}")

    # Controlla loghi mancanti
    img_dir = os.path.join(SITE_DIR, "images", "partners")
    missing = []
    for p in partners:
        if p["logo"]:
            logo_file = p["logo"].replace("images/partners/", "")
            full_path = os.path.join(img_dir, logo_file)
            if not os.path.exists(full_path):
                missing.append(logo_file)
    if missing:
        print(f"\n  ATTENZIONE — Loghi mancanti in images/partners/:")
        for m in missing:
            print(f"    - {m}")

    no_url = [p["name"] for p in partners if not p["url"]]
    if no_url:
        print(f"\n  ATTENZIONE — Partner senza sito web:")
        for n in no_url:
            print(f"    - {n}")


# ---------------------------------------------------------------------------
# ⛔⛔ CHIUSO A CHIAVE — il sito non si rigenera piu' dall'Excel (28/08/2026)
# ---------------------------------------------------------------------------
# Il titolare: «puoi bloccare l'aggiornamento automatico del sito dai file di
# excel. Bisogna fare in modo che il sito si aggiorni dal gestionale».
#
# ⚠️ **Non e' una precauzione teorica: questo script DISTRUGGE lavoro vero.**
# I file in `data/` portano ormai cose che nell'Excel non ci sono e non ci
# possono stare — le date dei singoli incontri, i luoghi con l'indirizzo, i
# loghi dei partner, le descrizioni, i nomi corretti dei progetti rinominati.
# Una rigenerazione le riportava tutte a stringa vuota **senza un errore e
# senza un avviso**: il sito si svuotava, e lo si scopriva guardandolo.
#
# ⛔ **Non si cancella lo script.** L'Excel resta la fonte storica del lavoro
# di sei anni, e il giorno in cui quei dati si travaseranno nel gestionale
# questo codice serve a rileggerlo. Chi lo lancia oggi, pero', quasi sempre non
# sa che cosa sta per perdere.
#
# ⇒ Per usarlo davvero: `--forza-so-cosa-sto-facendo`, e prima si fa una copia
#   di `data/`. Il nome dell'interruttore e' lungo apposta.
def _fermati_se_non_forzato():
    import sys as _sys
    if "--forza-so-cosa-sto-facendo" in _sys.argv:
        _sys.argv.remove("--forza-so-cosa-sto-facendo")
        print("\u26a0\ufe0f  Rigenerazione FORZATA: quel che era scritto a mano e a rischio.")
        return
    print(__doc__ or "")
    print("\u26d4 Questo script e' CHIUSO A CHIAVE dal 28/08/2026.")
    print()
    print("   Il sito non si rigenera piu' dall'Excel: i file in data/ contengono")
    print("   ormai dati che nell'Excel non esistono (appuntamenti, luoghi con")
    print("   indirizzo, loghi, descrizioni, nomi corretti), e rigenerare li")
    print("   cancella in silenzio.")
    print()
    print("   La strada nuova e' il gestionale: vedi docs/ e la memoria di")
    print("   progetto. Se DEVI davvero rigenerare, fai una copia di data/ e")
    print("   rilancia con  --forza-so-cosa-sto-facendo")
    raise SystemExit(2)

if __name__ == "__main__":
    _fermati_se_non_forzato()
    import argparse
    parser = argparse.ArgumentParser(description="Sincronizza partners.json")
    parser.add_argument("--partner-excel", help="Percorso Partner.xlsx")
    parser.add_argument("--projects-excel", help="Percorso Lista Progetti.xlsx")
    args = parser.parse_args()
    sync(args.partner_excel, args.projects_excel)
