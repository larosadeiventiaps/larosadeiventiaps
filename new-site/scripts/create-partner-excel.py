"""
Genera il file Partner.xlsx a partire da partners.json.
Uso una tantum per creare il file Excel iniziale.
"""

import json
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
JSON_PATH = os.path.join(SITE_DIR, "data", "partners.json")

ONEDRIVE_BASE = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Progetti",
)
OUTPUT_PATH = os.path.join(ONEDRIVE_BASE, "Partner.xlsx")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
    sys.exit(1)

with open(JSON_PATH, "r", encoding="utf-8") as f:
    partners = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Partner"

# Stili
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="E8630A", end_color="E8630A", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Intestazioni
headers = ["Nome", "Tipo", "Descrizione", "Logo", "Sito Web"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dati
for row_idx, partner in enumerate(partners, 2):
    logo_file = partner.get("logo", "").replace("images/partners/", "")
    ws.cell(row=row_idx, column=1, value=partner["name"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=partner.get("type", "")).border = thin_border
    ws.cell(row=row_idx, column=3, value=partner.get("description", "")).border = thin_border
    ws.cell(row=row_idx, column=4, value=logo_file).border = thin_border
    ws.cell(row=row_idx, column=5, value=partner.get("url", "")).border = thin_border

# Larghezze colonne
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 80
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 40

# Auto-filtro
ws.auto_filter.ref = f"A1:E{len(partners) + 1}"

# Salva
output = sys.argv[1] if len(sys.argv) > 1 else OUTPUT_PATH
os.makedirs(os.path.dirname(output), exist_ok=True)
wb.save(output)
print(f"File Excel creato: {output}")
print(f"  {len(partners)} partner inseriti")
