"""
Sincronizza gallery.json dal file Excel "gallery.xlsx".

Uso:
  python sync-gallery.py [<percorso-excel>]

Senza argomenti usa il file predefinito su OneDrive.

Colonne Excel:
  Titolo | Data | Luogo | Descrizione | Immagine
"""

import json
import os
import re
import sys
from datetime import datetime, date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(SITE_DIR, "data", "gallery.json")

DEFAULT_EXCEL = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Eventi",
    "gallery.xlsx",
)

MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}


def parse_date(val):
    """Converte un valore cella in una data ISO (YYYY-MM-DD)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    match = re.match(r"(\d{1,2})\s+(\w+)\s+(\d{4})", s)
    if match:
        day, month_name, year = match.groups()
        month = MESI_IT.get(month_name.lower())
        if month:
            return f"{year}-{month:02d}-{int(day):02d}"
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if match:
        day, month, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if match:
        return s
    return None


def sync(excel_path=None):
    """Legge l'Excel e genera gallery.json."""
    try:
        import openpyxl
    except ImportError:
        print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
        sys.exit(1)

    path = excel_path or DEFAULT_EXCEL
    if not os.path.exists(path):
        print(f"Errore: file non trovato: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    photos = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 1:
            continue

        title = row[0]
        if not title or str(title).strip() == "":
            continue
        title = str(title).strip()

        date_str = parse_date(row[1]) if len(row) > 1 else None
        location = str(row[2]).strip() if len(row) > 2 and row[2] else None
        description = str(row[3]).strip() if len(row) > 3 and row[3] else title
        img_val = row[4] if len(row) > 4 else None

        if img_val and str(img_val).strip():
            image = f"images/gallery/{str(img_val).strip()}"
        else:
            image = "images/logo.jpg"

        photo = {"title": title}
        if description:
            photo["description"] = description
        photo["image"] = image
        if date_str:
            photo["date"] = date_str
        if location:
            photo["location"] = location

        photos.append(photo)

    # Ordina per data decrescente
    photos.sort(key=lambda p: p.get("date", "0000"), reverse=True)

    # Salva JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(photos, f, ensure_ascii=False, indent=2)

    print(f"Sincronizzazione gallery completata!")
    print(f"  {len(photos)} foto")
    print(f"  Salvato in: {OUTPUT_PATH}")

    # Controlla immagini mancanti
    missing = []
    for p in photos:
        if p["image"] == "images/logo.jpg":
            continue
        full_path = os.path.join(SITE_DIR, p["image"])
        if not os.path.exists(full_path):
            missing.append(p["image"])
    if missing:
        print(f"\n  ATTENZIONE — Immagini mancanti:")
        for m in missing:
            print(f"    - {m}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    sync(path)
