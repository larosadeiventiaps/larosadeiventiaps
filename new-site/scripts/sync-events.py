"""
Sincronizza events.json dal file Excel "eventi.xlsx".

Uso:
  python sync-events.py [<percorso-excel>]

Senza argomenti usa il file predefinito su OneDrive.

Colonne Excel:
  Titolo | Data Inizio | Data Fine | Luogo | Descrizione | Immagine | Link
"""

import json
import os
import re
import sys
from datetime import datetime, date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(SITE_DIR, "data", "events.json")

DEFAULT_EXCEL = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Eventi",
    "eventi.xlsx",
)

MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}


def parse_date(val):
    """Converte un valore cella in una data ISO (YYYY-MM-DD)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    # Formato "7 luglio 2024"
    match = re.match(r"(\d{1,2})\s+(\w+)\s+(\d{4})", s)
    if match:
        day, month_name, year = match.groups()
        month = MESI_IT.get(month_name.lower())
        if month:
            return f"{year}-{month:02d}-{int(day):02d}"
    # Formato "26/11/2020"
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if match:
        day, month, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    # ISO
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if match:
        return s
    return None


def parse_time(val):
    """Converte un valore cella in formato HH:MM."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if not s:
        return None
    # Formato "15:30" o "15.30"
    match = re.match(r"(\d{1,2})[:\.](\d{2})", s)
    if match:
        h, m = match.groups()
        return f"{int(h):02d}:{m}"
    return None


def determine_status(date_str):
    """Determina se l'evento è futuro o passato."""
    if not date_str:
        return "passato"
    try:
        event_date = date.fromisoformat(date_str)
        return "futuro" if event_date >= date.today() else "passato"
    except ValueError:
        return "passato"


def sync(excel_path=None):
    """Legge l'Excel e genera events.json."""
    try:
        import openpyxl
    except ImportError:
        print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
        sys.exit(1)

    path = excel_path or DEFAULT_EXCEL
    if not os.path.exists(path):
        print(f"Errore: file non trovato: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    events = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 1:
            continue

        title = row[0]
        if not title or str(title).strip() == "":
            continue
        title = str(title).strip()

        start_str = parse_date(row[1]) if len(row) > 1 else None
        end_str = parse_date(row[2]) if len(row) > 2 else None
        location = str(row[3]).strip() if len(row) > 3 and row[3] else None
        description = str(row[4]).strip() if len(row) > 4 and row[4] else None
        img_val = row[5] if len(row) > 5 else None
        link = str(row[6]).strip() if len(row) > 6 and row[6] else None

        # Se non c'è data fine, usa data inizio
        if not end_str and start_str:
            end_str = start_str

        # Immagine
        if img_val and str(img_val).strip():
            image = f"images/events/{str(img_val).strip()}"
        else:
            image = "images/logo.jpg"

        status = determine_status(end_str or start_str)

        event = {"title": title}
        if start_str:
            event["startDate"] = start_str
        if end_str:
            event["endDate"] = end_str
        if location:
            event["location"] = location
        if description:
            event["description"] = description
        else:
            event["description"] = title
        event["image"] = image
        event["status"] = status
        if link:
            event["link"] = link

        events.append(event)

    # Ordina tutti per data decrescente (più recente prima)
    events.sort(key=lambda e: e.get("startDate", "0000"), reverse=True)

    # Salva JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)

    n_futuri = sum(1 for e in events if e["status"] == "futuro")
    n_passati = sum(1 for e in events if e["status"] == "passato")
    print(f"Sincronizzazione eventi completata!")
    print(f"  {len(events)} eventi ({n_futuri} futuri, {n_passati} passati)")
    print(f"  Salvato in: {OUTPUT_PATH}")

    # Controlla immagini mancanti
    missing = []
    for e in events:
        if e["image"] == "images/logo.jpg":
            continue
        full_path = os.path.join(SITE_DIR, e["image"])
        if not os.path.exists(full_path):
            missing.append(e["image"])
    if missing:
        print(f"\n  ATTENZIONE — Immagini mancanti:")
        for m in missing:
            print(f"    - {m}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    sync(path)
