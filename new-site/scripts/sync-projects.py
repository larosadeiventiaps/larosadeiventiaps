"""
Sincronizza projects.json dal file Excel "Lista Progetti.xlsx".

Uso:
  python sync-projects.py <percorso-excel>

Se non viene passato un percorso, usa il file predefinito su OneDrive.

La colonna "Immagini" dell'Excel deve contenere il nome del file immagine
(es. "genitori-a-confronto.jpg"). Le immagini vanno messe nella cartella
new-site/images/projects/.
"""

import json
import os
import re
import sys
from datetime import datetime, date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(SITE_DIR, "data", "projects.json")

DEFAULT_EXCEL = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Progetti",
    "Lista Progetti.xlsx",
)

# Mappa colonne Excel -> campi JSON
COL_MAP = {
    0: "startDate",
    1: "endDate",
    2: "title",
    3: "description",
    4: "incontri",
    5: "oreIncontro",
    6: "partecipanti",
    7: "educatori",
    8: "volontari",
    9: "sponsor",
    10: "professionisti",
    11: "collaboratori",
    12: "ore",
    13: "image",
}

MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}


def parse_date(val):
    """Converte un valore cella in una data ISO (YYYY-MM-DD)."""
    if val is None:
        return None

    # Già un oggetto datetime/date
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()

    s = str(val).strip()
    if not s:
        return None

    # Formato "7 luglio 2024" o "21 marzo 2025"
    match = re.match(r"(\d{1,2})\s+(\w+)\s+(\d{4})", s)
    if match:
        day, month_name, year = match.groups()
        month = MESI_IT.get(month_name.lower())
        if month:
            return f"{year}-{month:02d}-{int(day):02d}"

    # Formato "gennaio 2023" (senza giorno)
    match = re.match(r"(\w+)\s+(\d{4})", s)
    if match:
        month_name, year = match.groups()
        month = MESI_IT.get(month_name.lower())
        if month:
            return f"{year}-{month:02d}-01"

    # Formato "26/11/2020"
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if match:
        day, month, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"

    # Anno solo "2023"
    match = re.match(r"^(\d{4})$", s)
    if match:
        return f"{s}-01-01"

    return None


def determine_status(start_str, end_str):
    """Determina lo stato del progetto in base alle date."""
    today = date.today()

    start = None
    end = None
    if start_str:
        try:
            start = date.fromisoformat(start_str)
        except ValueError:
            pass
    if end_str:
        try:
            end = date.fromisoformat(end_str)
        except ValueError:
            pass

    if start and start > today:
        return "futuro"
    if end and end < today:
        return "passato"
    if start and start <= today:
        return "in_corso"

    return "passato"


def to_int(val):
    """Converte in intero se possibile."""
    if val is None:
        return None
    try:
        n = int(float(val))
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def to_float(val):
    """Converte in float se possibile."""
    if val is None:
        return None
    try:
        n = float(val)
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def build_description(project):
    """Genera una descrizione dal titolo e dai collaboratori/sponsor."""
    parts = []
    title = project.get("title", "")

    if project.get("collaboratori"):
        parts.append(f"In collaborazione con {project['collaboratori']}.")
    if project.get("sponsor"):
        parts.append(f"Con il sostegno di {project['sponsor']}.")
    if project.get("incontri") and project.get("ore"):
        parts.append(f"{project['incontri']} incontri per un totale di {project['ore']} ore.")
    elif project.get("incontri"):
        parts.append(f"{project['incontri']} incontri.")
    if project.get("partecipanti"):
        parts.append(f"{project['partecipanti']} partecipanti.")

    return " ".join(parts) if parts else title


def sync(excel_path):
    """Legge l'Excel e genera projects.json."""
    try:
        import openpyxl
    except ImportError:
        print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(excel_path):
        print(f"Errore: file non trovato: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    projects = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Salta righe vuote o righe di riepilogo (colonna 14/15 con "Totale...")
        if not row or len(row) < 3:
            continue

        title = row[2]
        if not title or str(title).strip() == "":
            continue

        title = str(title).strip()

        # Salta se sembra una riga di riepilogo
        if title.startswith("Totale") or title.startswith("Nr."):
            continue

        start_raw = parse_date(row[0])
        end_raw = parse_date(row[1]) if len(row) > 1 else None

        # Se non c'è data fine, usa data inizio
        if not end_raw and start_raw:
            end_raw = start_raw

        status = determine_status(start_raw, end_raw)

        # Colonna 3 = Descrizione Progetto (dall'Excel)
        excel_description = str(row[3]).strip() if len(row) > 3 and row[3] else None

        # Colonna 13 = Immagine
        img_val = row[13] if len(row) > 13 else None
        if img_val and str(img_val).strip() and str(img_val).strip().lower() != "x":
            image = f"images/projects/{str(img_val).strip()}"
        else:
            image = "images/logo.jpg"

        project = {"title": title}

        # Campi numerici (colonne 4-12, la 3 è descrizione)
        incontri = to_int(row[4]) if len(row) > 4 else None
        ore_incontro = to_float(row[5]) if len(row) > 5 else None
        partecipanti = to_int(row[6]) if len(row) > 6 else None
        educatori = to_int(row[7]) if len(row) > 7 else None
        volontari = to_int(row[8]) if len(row) > 8 else None
        sponsor = str(row[9]).strip() if len(row) > 9 and row[9] else None
        professionisti = to_int(row[10]) if len(row) > 10 else None
        collaboratori = str(row[11]).strip() if len(row) > 11 and row[11] else None
        ore_totali = to_float(row[12]) if len(row) > 12 else None

        project["image"] = image
        if start_raw:
            project["startDate"] = start_raw
        if end_raw:
            project["endDate"] = end_raw
        project["status"] = status

        if incontri:
            project["incontri"] = incontri
        if ore_totali and ore_totali > 0:
            project["ore"] = ore_totali
        if partecipanti:
            project["partecipanti"] = partecipanti
        if educatori:
            project["educatori"] = educatori
        if volontari:
            project["volontari"] = volontari
        if professionisti:
            project["professionisti"] = professionisti
        if sponsor:
            project["sponsor"] = sponsor
        if collaboratori:
            project["collaboratori"] = collaboratori

        # Usa la descrizione dall'Excel se presente, altrimenti genera automatica
        if excel_description:
            project["description"] = excel_description
        else:
            project["description"] = build_description(project)

        projects.append(project)

    # In corso e futuri: dal più vicino al più lontano (crescente)
    # Passati: dal più recente al più vecchio (decrescente)
    in_corso = sorted([p for p in projects if p["status"] == "in_corso"],
                      key=lambda p: p.get("startDate", "9999"))
    futuri = sorted([p for p in projects if p["status"] == "futuro"],
                    key=lambda p: p.get("startDate", "9999"))
    passati = sorted([p for p in projects if p["status"] == "passato"],
                     key=lambda p: p.get("startDate", "0000"), reverse=True)
    projects = in_corso + futuri + passati

    # Salva JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)

    # Riepilogo
    tot = len(projects)
    in_corso = sum(1 for p in projects if p["status"] == "in_corso")
    futuro = sum(1 for p in projects if p["status"] == "futuro")
    passato = sum(1 for p in projects if p["status"] == "passato")
    tot_ore = sum(p.get("ore", 0) for p in projects)
    tot_part = sum(p.get("partecipanti", 0) for p in projects)
    tot_incontri = sum(p.get("incontri", 0) for p in projects)

    print(f"Sincronizzazione completata!")
    print(f"  {tot} progetti ({in_corso} in corso, {futuro} futuri, {passato} passati)")
    print(f"  {tot_incontri} incontri, {tot_ore} ore, {tot_part} partecipanti")
    print(f"  Salvato in: {OUTPUT_PATH}")

    # Controlla immagini mancanti
    missing = []
    for p in projects:
        img_path = p["image"]
        if img_path == "images/logo.jpg":
            continue  # fallback al logo, sempre presente
        full_path = os.path.join(SITE_DIR, img_path)
        if not os.path.exists(full_path):
            missing.append(img_path)
    if missing:
        print(f"\n  ATTENZIONE — Immagini mancanti in images/projects/:")
        for m in missing:
            print(f"    - {m}")


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    sync(excel_path)
