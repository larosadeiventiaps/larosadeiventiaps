"""
Sincronizza partners.json dal file Excel "Partner.xlsx" e/o dalla lista progetti.

Uso:
  python sync-partners.py [--partner-excel <percorso>] [--projects-excel <percorso>]

Senza argomenti usa i file predefiniti su OneDrive.

Il file Partner.xlsx ha le colonne:
  Nome | Tipo | Descrizione | Logo | Sito Web

Il file Lista Progetti.xlsx viene usato per estrarre nomi partner
dai campi sponsor e collaboratori. Eventuali nuovi partner trovati
vengono aggiunti con dati placeholder.
"""

import json
import os
import re
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(SITE_DIR, "data", "partners.json")

ONEDRIVE_BASE = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Progetti",
)

DEFAULT_PARTNER_EXCEL = os.path.join(ONEDRIVE_BASE, "Partner.xlsx")
DEFAULT_PROJECTS_EXCEL = os.path.join(ONEDRIVE_BASE, "Lista Progetti.xlsx")


# Mappa abbreviazioni/varianti usate nell'Excel progetti → nome completo
# nel Partner.xlsx. I nomi qui sotto vengono ignorati quando appaiono
# nell'Excel progetti perché sono già coperti dal partner corretto.
ALIASES = {
    "comune bar": "Comune di Bagno a Ripoli",
    "bar": "Comune di Bagno a Ripoli",
    "fondazione cr": "Fondazione CR Firenze",
    "cri": "Croce Rossa Italiana",
    "biblioteca": "Biblioteca di Bagno a Ripoli",
    "scuola redi": "Scuola Francesco Redi",
    "elsa morante": "Istituto Elsa Morante",
    "cdp grassina": "Casa del Popolo Grassina",
    "cdp balatro": "Casa del Popolo Balatro",
    "contrada alfiere": "Contrada dell'Alfiere",
    "becare": "BE Care S.r.l.",
    "mise antella": "Misericordia di Antella",
    "mise ponte a ema": "Misericordia di Ponte a Ema",
    "mise badia": "Misericordia di Badia a Ripoli",
    "fpgrassina": "Fratellanza Popolare Grassina",
    "croce d'oro ponte a ema": "Croce d'Oro Ponte a Ema",
    "l'apiario": "Apicoltura San Martino",
}

# Stringhe composte nell'Excel che vanno espanse in più partner singoli
COMPOSITE_EXPANSIONS = {
    "cri mise antella fpgrassina mise ponte a ema mise badia": [
        "Croce Rossa Italiana",
        "Misericordia di Antella",
        "Fratellanza Popolare Grassina",
        "Misericordia di Ponte a Ema",
        "Misericordia di Badia a Ripoli",
    ],
    "sds firenze sud-est e comune bar": [
        "SdS Firenze Sud-Est",
        "Comune di Bagno a Ripoli",
    ],
    "comune bar il teatro dell'inutile e l'apiario": [
        "Comune di Bagno a Ripoli",
        "Il Teatro dell'Inutile",
        "Apicoltura San Martino",
    ],
    "la lanterna scuola redi": [
        "La Lanterna",
        "Scuola Francesco Redi",
    ],
    "scuola redi fondazione claudio ciai": [
        "Scuola Francesco Redi",
        "Fondazione Claudio Ciai",
    ],
    "becare e cri": [
        "BE Care S.r.l.",
        "Croce Rossa Italiana",
    ],
    "auser legambiente": [
        "Auser",
        "Legambiente",
    ],
    "consorzio blu ancora": [
        "Consorzio Blu",
    ],
    "elsa morante e croce d'oro ponte a ema": [
        "Istituto Elsa Morante",
        "Croce d'Oro Ponte a Ema",
    ],
}


def normalize_name(name):
    """Normalizza un nome partner per confronto."""
    return re.sub(r"\s+", " ", name.strip().lower())


def resolve_partner_names(raw):
    """Risolve una stringa dall'Excel progetti in nomi partner reali."""
    if not raw or not str(raw).strip():
        return []
    s = str(raw).strip()
    key = normalize_name(s)

    # Controlla se è una stringa composita nota
    if key in COMPOSITE_EXPANSIONS:
        return COMPOSITE_EXPANSIONS[key]

    # Controlla se è un alias noto
    if key in ALIASES:
        return [ALIASES[key]]

    # Prova a splittare su separatori semplici (" e ", ", ")
    parts = re.split(r"\s+e\s+|,\s*|/\s*", s)
    resolved = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        pk = normalize_name(p)
        if pk in ALIASES:
            resolved.append(ALIASES[pk])
        else:
            resolved.append(p)
    return resolved


def read_partner_excel(path):
    """Legge Partner.xlsx e restituisce una lista di dict."""
    try:
        import openpyxl
    except ImportError:
        print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(path):
        print(f"Partner.xlsx non trovato: {path}")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    partners = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue

        name = str(row[0]).strip()
        partner = {
            "name": name,
            "type": str(row[1]).strip() if len(row) > 1 and row[1] else "associazione",
            "description": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "logo": "",
            "url": "",
        }

        if len(row) > 3 and row[3] and str(row[3]).strip():
            logo_file = str(row[3]).strip()
            partner["logo"] = f"images/partners/{logo_file}"

        if len(row) > 4 and row[4] and str(row[4]).strip():
            partner["url"] = str(row[4]).strip()

        partners.append(partner)

    return partners


def extract_from_projects(path):
    """Estrae nomi partner unici dall'Excel dei progetti."""
    try:
        import openpyxl
    except ImportError:
        return set()

    if not os.path.exists(path):
        print(f"Lista Progetti.xlsx non trovato: {path}")
        return set()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        title = row[2]
        if not title or str(title).strip() in ("", "Totale", "Nr."):
            continue

        # Colonna 8 = sponsor, colonna 10 = collaboratori
        sponsor = row[8] if len(row) > 8 else None
        collaboratori = row[10] if len(row) > 10 else None

        for raw in [sponsor, collaboratori]:
            for name in resolve_partner_names(raw):
                names.add(name)

    return names


def sync(partner_excel=None, projects_excel=None):
    """Sincronizza partners.json."""
    partner_path = partner_excel or DEFAULT_PARTNER_EXCEL
    projects_path = projects_excel or DEFAULT_PROJECTS_EXCEL

    # 1. Leggi partner da Excel dedicato
    partners = read_partner_excel(partner_path)
    known_names = {normalize_name(p["name"]) for p in partners}

    # 2. Estrai nomi da Excel progetti
    project_names = extract_from_projects(projects_path)

    # 3. Aggiungi partner nuovi (non già presenti)
    added = 0
    for name in sorted(project_names):
        if normalize_name(name) not in known_names:
            partners.append({
                "name": name,
                "type": "associazione",
                "description": "",
                "logo": "",
                "url": "",
            })
            known_names.add(normalize_name(name))
            added += 1

    # 4. Ordina per nome
    partners.sort(key=lambda p: p["name"].lower())

    # 5. Salva JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(partners, f, ensure_ascii=False, indent=2)

    print(f"Sincronizzazione partner completata!")
    print(f"  {len(partners)} partner totali ({added} nuovi da progetti)")
    print(f"  Salvato in: {OUTPUT_PATH}")

    # Controlla loghi mancanti
    img_dir = os.path.join(SITE_DIR, "images", "partners")
    missing = []
    for p in partners:
        if p["logo"]:
            logo_file = p["logo"].replace("images/partners/", "")
            full_path = os.path.join(img_dir, logo_file)
            if not os.path.exists(full_path):
                missing.append(logo_file)
    if missing:
        print(f"\n  ATTENZIONE — Loghi mancanti in images/partners/:")
        for m in missing:
            print(f"    - {m}")

    no_url = [p["name"] for p in partners if not p["url"]]
    if no_url:
        print(f"\n  ATTENZIONE — Partner senza sito web:")
        for n in no_url:
            print(f"    - {n}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Sincronizza partners.json")
    parser.add_argument("--partner-excel", help="Percorso Partner.xlsx")
    parser.add_argument("--projects-excel", help="Percorso Lista Progetti.xlsx")
    args = parser.parse_args()
    sync(args.partner_excel, args.projects_excel)
