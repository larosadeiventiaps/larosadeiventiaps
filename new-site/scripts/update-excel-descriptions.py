"""
Aggiorna Lista Progetti.xlsx con descrizioni e nomi immagini.
Uso una tantum per popolare i campi vuoti.
"""

import os
import sys
import shutil

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SITE_DIR = os.path.dirname(SCRIPT_DIR)

ONEDRIVE_BASE = os.path.join(
    os.path.expanduser("~"),
    "La Rosa dei Venti Aps",
    "La Rosa dei Venti Aps - Documenti",
    "Progetti",
)
EXCEL_PATH = os.path.join(ONEDRIVE_BASE, "Lista Progetti.xlsx")
IMAGES_DEST = os.path.join(ONEDRIVE_BASE, "Immagini Progetto per Sito")
IMAGES_SRC = os.path.join(SITE_DIR, "images", "projects")

try:
    import openpyxl
except ImportError:
    print("Errore: openpyxl non installato. Esegui: pip install openpyxl")
    sys.exit(1)

# Mappa titolo progetto -> immagine suggerita
IMAGE_MAP = {
    "Follow me...Ad uno s-passo dalla cultura": "ad-uno-spasso-dalla-cultura.jpg",
    "Centro Estivo": "centro-estivo.jpg",
    "Cinema": "cinema-movie-time.jpeg",
    "Conversazioni con il pc": "corso-informatica.jpg",
    "Genitori a confronto/conforto": "genitori-a-confronto.jpg",
    "Giochiamo con l'inglese": "giocare-con-inglese.jpg",
    "Un gioco per tutti": "gioco-libero.jpg",
    "Pomeriggio con i lego": "mattoncini-lego.jpg",
    "Percorso Didattico Naturalistico parco Villa Mondeggi": "fattorie-didattiche.jpg",
    "Marciapiede Didattico": "bagno-a-ripoli.jpg",
    "Dall'orto con amore": "corso-riciclo.jpg",
    "Passeggiate Ripolesi": "insieme-per-orientarsi.jpg",
    "Colori in libertà": "corso-teatro.jpg",
    "Un pennello per amico": "corso-teatro.jpg",
    "Espressione corporea": "musica-e-movimento.jpeg",
    "Settimana dei Diritti": "bagno-a-ripoli-rontini.jpg",
    "Prim'olio": "la-tommasina.jpg",
    "Il Lavandeto": "fattorie-didattiche.jpg",
    "Una coccola in corsia": "educazione-casalinga.jpg",
    "Orto nel bosco Orti sociali": "corso-riciclo.jpg",
    "Frittelle": "la-tommasina.jpg",
}

# Descrizioni per progetto
DESCRIPTIONS = {
    "Un gioco per tutti": "Progetto di socializzazione attraverso il gioco per bambini e ragazzi con disabilità, favorendo l'inclusione e il divertimento condiviso.",
    "Una coccola in corsia": "Iniziativa di volontariato per portare comfort e vicinanza ai pazienti ricoverati attraverso piccoli gesti di cura e attenzione.",
    "Dall'orto con amore": "Laboratorio di orticultura educativa per imparare a coltivare ortaggi e piante, promuovendo il contatto con la natura e il lavoro di squadra.",
    "Diritto parità di genere": "Incontro di sensibilizzazione sulla parità di genere e i diritti delle persone con disabilità, in collaborazione con il Comune di Bagno a Ripoli.",
    "Passeggiate Ripolesi": "Escursioni guidate alla scoperta del territorio di Bagno a Ripoli, tra sentieri, borghi e paesaggi collinari.",
    "Giochiamo con l'inglese": "Laboratorio ludico-didattico di lingua inglese per ragazzi, con giochi, conversazione e attività interattive per imparare divertendosi.",
    "Marciapiede Didattico": "Progetto di educazione stradale per le scuole con percorsi pratici, in collaborazione con Croce Rossa, Misericordie e Fratellanza Popolare.",
    "Conversazioni con il pc": "Corso di alfabetizzazione informatica per imparare ad usare il computer, navigare in internet e utilizzare i principali strumenti digitali.",
    "Follow me...Ad uno s-passo dalla cultura": "Percorso di uscite culturali guidate a Firenze e dintorni per scoprire monumenti, musei e bellezze del patrimonio artistico toscano.",
    "Cinema": "Ciclo di proiezioni cinematografiche seguite da discussione di gruppo, per condividere emozioni e riflessioni attraverso il grande schermo.",
    "Orto nel bosco Orti sociali": "Progetto di orti sociali nel verde del bosco, per coltivare insieme ortaggi e piante promuovendo la socializzazione all'aria aperta.",
    "Frittelle": "Laboratorio di cucina e convivialità legato alle tradizioni locali, con preparazione di frittelle e momenti di festa insieme.",
    "Pomeriggio con i lego": "Pomeriggio creativo con i mattoncini Lego per stimolare la fantasia, la manualità e la collaborazione tra i partecipanti.",
    "Genitori a confronto/conforto": "Gruppo di sostegno e confronto tra genitori su tematiche legate alla genitorialità, alla disabilità e alla crescita dei figli.",
    "Un pennello per amico": "Laboratorio di pittura e arti visive dove esprimere la propria creatività attraverso colori, pennelli e tecniche artistiche diverse.",
    "Espressione corporea": "Laboratorio di movimento e espressione corporea per sviluppare consapevolezza del corpo, coordinazione e capacità comunicative.",
    "Centro Estivo": "Settimana di attività estive per ragazzi con giochi, laboratori, uscite e momenti di socializzazione durante l'estate.",
    "Percorso Didattico Naturalistico parco Villa Mondeggi": "Escursione naturalistica al parco di Villa Mondeggi per scoprire la biodiversità, le piante e gli animali del territorio.",
    "Aiutiamo l'ambiente": "Giornata di volontariato ambientale per la pulizia e la cura degli spazi verdi, in collaborazione con Angeli del Bello.",
    "Uscita con il Trekking BaR": "Escursione guidata con il Gruppo Trekking di Bagno a Ripoli alla scoperta dei sentieri e dei paesaggi del territorio ripolese.",
    "Un dolce ronzio": "Laboratorio di apicoltura didattica per scoprire il mondo delle api, la produzione del miele e l'importanza degli impollinatori.",
    "Il Lavandeto": "Visita al lavandeto con laboratorio esperienziale tra piante aromatiche, teatro e apiario, alla scoperta della natura e delle tradizioni.",
    "Settimana dei Diritti": "Settimana di eventi e incontri nelle scuole sui diritti delle persone con disabilità, con professionisti e testimonianze dirette.",
    "Prim'olio": "Giornata dedicata alla raccolta delle olive e alla produzione dell'olio, per vivere insieme le tradizioni agricole del territorio.",
    "Colori in libertà": "Laboratorio artistico di pittura libera e creativa, per esprimersi attraverso i colori senza vincoli e in totale libertà.",
    "Rifirma sulla disabilità": "Incontro informativo sulla nuova riforma della disabilità e il Progetto di Vita, con esperti e operatori del settore.",
    "Volontmusic": "Contest musicale tra band emergenti delle associazioni di volontariato per promuovere la solidarietà attraverso la musica.",
}

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[wb.sheetnames[0]]

updated_desc = 0
updated_img = 0

for row_idx in range(2, ws.max_row + 1):
    title_cell = ws.cell(row=row_idx, column=3)
    if not title_cell.value or str(title_cell.value).strip() == "":
        continue

    title = str(title_cell.value).strip()

    # Salta righe riepilogo
    if title.startswith("Totale") or title.startswith("Nr."):
        continue

    # Aggiorna descrizione (colonna 4) se vuota
    desc_cell = ws.cell(row=row_idx, column=4)
    if not desc_cell.value or str(desc_cell.value).strip() == "":
        if title in DESCRIPTIONS:
            desc_cell.value = DESCRIPTIONS[title]
            updated_desc += 1

    # Aggiorna immagine (colonna 14) se vuota
    img_cell = ws.cell(row=row_idx, column=14)
    if not img_cell.value or str(img_cell.value).strip() == "":
        if title in IMAGE_MAP:
            img_cell.value = IMAGE_MAP[title]
            updated_img += 1

wb.save(EXCEL_PATH)
print(f"Excel aggiornato: {EXCEL_PATH}")
print(f"  {updated_desc} descrizioni aggiunte")
print(f"  {updated_img} immagini assegnate")

# Copia immagini nella cartella OneDrive
copied = 0
for img_file in os.listdir(IMAGES_SRC):
    if img_file == "placeholder.svg":
        continue
    src = os.path.join(IMAGES_SRC, img_file)
    dst = os.path.join(IMAGES_DEST, img_file)
    if os.path.isfile(src) and not os.path.exists(dst):
        shutil.copy2(src, dst)
        copied += 1

print(f"  {copied} immagini copiate in: {IMAGES_DEST}")
